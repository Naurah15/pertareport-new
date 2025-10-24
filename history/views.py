import urllib.parse
import os
import openpyxl
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from report.models import Laporan
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from report.models import Laporan
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from openpyxl.drawing.image import Image as XLImage
from django.conf import settings
from openpyxl.utils import get_column_letter
from reportlab.lib.units import inch
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from datetime import datetime
from django.utils import timezone
from django.utils.dateparse import parse_date
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.colors import HexColor
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib import colors


@login_required
def history_list(request):
    user = request.user
    
    # Get base queryset based on user permissions
    if user.username == 'admin' and user.check_password('Mimin1234%'):
        laporan_queryset = Laporan.objects.all()
    else:
        laporan_queryset = Laporan.objects.filter(
            nama_team_support__iexact=user.username
        )
    
    # Apply date filters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date:
        try:
            start_date_parsed = parse_date(start_date)
            if start_date_parsed:
                laporan_queryset = laporan_queryset.filter(
                    tanggal_proses__date__gte=start_date_parsed
                )
        except ValueError:
            pass
    
    if end_date:
        try:
            end_date_parsed = parse_date(end_date)
            if end_date_parsed:
                laporan_queryset = laporan_queryset.filter(
                    tanggal_proses__date__lte=end_date_parsed
                )
        except ValueError:
            pass
    
    # Order by name first, then by date
    laporan_list = laporan_queryset.order_by('nama_team_support', '-tanggal_proses')

    return render(request, 'history_list.html', {
        'laporan_list': laporan_list
    })


@login_required
def download_laporan_excel(request, pk):
    laporan = get_object_or_404(Laporan, pk=pk)
    kegiatan_list = laporan.kegiatan_list.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan"

    # ==== Styling dasar ====
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="003876", end_color="003876", fill_type="solid")
    title_font = Font(bold=True, size=14, color="003876")
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ==== Judul ====
    ws.merge_cells('A1:C1')
    cell = ws['A1']
    cell.value = f"LAPORAN {laporan.no_document}"
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center")

    ws.append([])
    ws.append(["No Document", laporan.no_document])
    coord = (laporan.lokasi or "").strip()
    ws.append(["Lokasi", coord])
    if laporan.spbu:
        ws.append(["SPBU", f"{laporan.spbu.kode} - {laporan.spbu.nama}"])
        if laporan.spbu.alamat:
            ws.append(["Alamat SPBU", laporan.spbu.alamat])
    else:
        ws.append(["SPBU", "Tidak ada"])
    ws["B4"].hyperlink = f"https://www.google.com/maps/search/?api=1&query={urllib.parse.quote(coord)}"
    ws["B4"].style = "Hyperlink"

    ws.append(["Nama Team Support", laporan.nama_team_support])
    ws.append(["Tanggal Proses", laporan.tanggal_proses.strftime("%d-%m-%Y %H:%M")])
    ws.append([])

    # ==== Header tabel ====
    header_row = ws.max_row + 1
    headers = ["Jenis Kegiatan", "Remark", "Foto"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # ==== Data & gambar ====
    row_num = ws.max_row + 1
    for kegiatan in kegiatan_list:
        ws.append([kegiatan.get_kegiatan_display_name(), kegiatan.remark, ""])
        
        # Set tinggi baris dulu
        ws.row_dimensions[row_num].height = 155
        
        # Set alignment dan border untuk kolom A dan B
        ws.cell(row=row_num, column=1).alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        ws.cell(row=row_num, column=1).border = border_style
        ws.cell(row=row_num, column=2).alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
        ws.cell(row=row_num, column=2).border = border_style
        
        foto_col = 3  # mulai dari kolom C

        # Foto utama
        if kegiatan.foto:
            img_path = os.path.join(settings.MEDIA_ROOT, str(kegiatan.foto))
            if os.path.exists(img_path):
                img = XLImage(img_path)
                img.width = 280
                img.height = 200
                img.anchor = f"{get_column_letter(foto_col)}{row_num}"
                ws.add_image(img)
                ws.column_dimensions[get_column_letter(foto_col)].width = 42
                foto_col += 1

        # Foto tambahan
        for foto in kegiatan.foto_list.all():
            img_path = os.path.join(settings.MEDIA_ROOT, str(foto.foto))
            if os.path.exists(img_path):
                img = XLImage(img_path)
                img.width = 280
                img.height = 200
                col_letter = get_column_letter(foto_col)
                img.anchor = f"{col_letter}{row_num}"
                ws.add_image(img)
                ws.column_dimensions[col_letter].width = 42
                foto_col += 1

        # Apply border untuk semua kolom foto
        for c in range(3, foto_col):
            ws.cell(row=row_num, column=c).border = border_style
            ws.cell(row=row_num, column=c).alignment = Alignment(vertical="center", horizontal="center")

        row_num += 1

    # ==== Kolom teks utama ====
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="Laporan_{laporan.no_document}.xlsx"'
    wb.save(response)
    return response


@login_required
def download_laporan_pdf(request, pk):
    laporan = get_object_or_404(Laporan, pk=pk)
    kegiatan_list = laporan.kegiatan_list.all()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Laporan_{laporan.no_document}.pdf"'

    p = canvas.Canvas(response, pagesize=A4)
    page_width, page_height = A4
    
    # Header pertama
    y = 800
    p.setFillColor(HexColor('#003876'))
    p.rect(40, y-5, page_width-80, 25, fill=1, stroke=0)
    p.setFont("Helvetica-Bold", 16)
    p.setFillColor(HexColor('#FFFFFF'))
    p.drawCentredString(page_width/2, y, f"LAPORAN: {laporan.no_document}")
    y -= 40

    # Separator line
    p.setStrokeColor(HexColor('#E31E24'))
    p.setLineWidth(2)
    p.line(50, y, page_width - 50, y)
    y -= 25

    # Info dasar dengan layout lebih rapi
    p.setFont("Helvetica-Bold", 11)
    p.setFillColor(HexColor('#003876'))
    p.drawString(50, y, "Lokasi:")
    p.setFont("Helvetica", 10)
    p.setFillColor(HexColor('#0066CC'))
    coord = laporan.lokasi or ""
    p.linkURL(f"https://www.google.com/maps/search/?api=1&query={urllib.parse.quote(coord)}", 
            (150, y-2, 500, y+12))
    p.drawString(150, y, coord)
    y -= 18

    # SPBU info
    p.setFont("Helvetica-Bold", 11)
    p.setFillColor(HexColor('#003876'))
    p.drawString(50, y, "SPBU:")
    p.setFont("Helvetica", 10)
    p.setFillColor(HexColor('#000000'))
    if laporan.spbu:
        spbu_text = f"{laporan.spbu.kode} - {laporan.spbu.nama}"
        p.drawString(150, y, spbu_text)
        y -= 18
        if laporan.spbu.alamat:
            p.setFont("Helvetica-Bold", 11)
            p.setFillColor(HexColor('#003876'))
            p.drawString(50, y, "Alamat:")
            p.setFont("Helvetica", 10)
            p.setFillColor(HexColor('#000000'))
            p.drawString(150, y, laporan.spbu.alamat)
            y -= 18
    else:
        p.drawString(150, y, "Tidak ada")
        y -= 18

    # Team Support
    p.setFont("Helvetica-Bold", 11)
    p.setFillColor(HexColor('#003876'))
    p.drawString(50, y, "Team Support:")
    p.setFont("Helvetica", 10)
    p.setFillColor(HexColor('#000000'))
    p.drawString(150, y, laporan.nama_team_support)
    y -= 18
    
    # Tanggal
    p.setFont("Helvetica-Bold", 11)
    p.setFillColor(HexColor('#003876'))
    p.drawString(50, y, "Tanggal:")
    p.setFont("Helvetica", 10)
    p.setFillColor(HexColor('#000000'))
    p.drawString(150, y, laporan.tanggal_proses.strftime('%d-%m-%Y %H:%M'))
    y -= 35

    # Separator line sebelum kegiatan
    p.setStrokeColor(HexColor('#CCCCCC'))
    p.setLineWidth(1)
    p.line(50, y, page_width - 50, y)
    y -= 25

    for kegiatan in kegiatan_list:
        if y < 280:
            p.showPage()
            y = 800

        # Background untuk kegiatan
        p.setFillColor(HexColor('#F0F4F8'))
        p.rect(45, y-18, page_width-90, 23, fill=1, stroke=0)

        # Kegiatan header
        p.setFont("Helvetica-Bold", 12)
        p.setFillColor(HexColor('#003876'))
        p.drawString(55, y-12, f"• {kegiatan.get_kegiatan_display_name()}")
        y -= 25
        
        # Remark
        p.setFont("Helvetica", 10)
        p.setFillColor(HexColor('#000000'))
        remark_text = f"Remark: {kegiatan.remark}"
        p.drawString(65, y, remark_text)
        y -= 22

        # Foto utama dari KegiatanLaporan
        if kegiatan.foto:
            img_path = os.path.join(settings.MEDIA_ROOT, str(kegiatan.foto))
            if os.path.exists(img_path):
                if y < 195:
                    p.showPage()
                    y = 800
                p.drawImage(img_path, 65, y-165, width=240, height=165, preserveAspectRatio=True, mask='auto')
                y -= 175

        # Foto tambahan dari KegiatanFoto
        for foto in kegiatan.foto_list.all():
            img_path = os.path.join(settings.MEDIA_ROOT, str(foto.foto))
            if os.path.exists(img_path):
                if y < 195:
                    p.showPage()
                    y = 800
                p.drawImage(img_path, 65, y-165, width=240, height=165, preserveAspectRatio=True, mask='auto')
                y -= 175

        y -= 20

    p.showPage()
    p.save()
    return response


@login_required
def bulk_download_excel(request):
    user = request.user
    
    # Get filtered queryset
    if user.username == 'admin' and user.check_password('Mimin1234%'):
        laporan_queryset = Laporan.objects.all()
    else:
        laporan_queryset = Laporan.objects.filter(
            nama_team_support__iexact=user.username
        )
    
    # Apply date filters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date:
        try:
            start_date_parsed = parse_date(start_date)
            if start_date_parsed:
                laporan_queryset = laporan_queryset.filter(
                    tanggal_proses__date__gte=start_date_parsed
                )
        except ValueError:
            pass
    
    if end_date:
        try:
            end_date_parsed = parse_date(end_date)
            if end_date_parsed:
                laporan_queryset = laporan_queryset.filter(
                    tanggal_proses__date__lte=end_date_parsed
                )
        except ValueError:
            pass
    
    laporan_list = laporan_queryset.order_by('nama_team_support', '-tanggal_proses')
    
    if not laporan_list:
        response = HttpResponse("Tidak ada laporan untuk diunduh")
        return response

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan Bulk"

    # Styling
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="003876", end_color="003876", fill_type="solid")
    title_font = Font(bold=True, size=16, color="E31E24")
    section_font = Font(bold=True, size=14, color="003876")
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Main title
    ws.merge_cells('A1:H1')
    cell = ws['A1']
    cell.value = "LAPORAN KEGIATAN PERTAMINA"
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center")
    
    # Date range info
    if start_date and end_date:
        ws.merge_cells('A2:H2')
        cell = ws['A2']
        cell.value = f"Periode: {start_date} s/d {end_date}"
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center")
        current_row = 4
    elif start_date:
        ws.merge_cells('A2:H2')
        cell = ws['A2']
        cell.value = f"Dari Tanggal: {start_date}"
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center")
        current_row = 4
    elif end_date:
        ws.merge_cells('A2:H2')
        cell = ws['A2']
        cell.value = f"Sampai Tanggal: {end_date}"
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center")
        current_row = 4
    else:
        current_row = 3

    # Group by team support
    grouped_laporan = {}
    for laporan in laporan_list:
        if laporan.nama_team_support not in grouped_laporan:
            grouped_laporan[laporan.nama_team_support] = []
        grouped_laporan[laporan.nama_team_support].append(laporan)

    for team_name, team_laporan in grouped_laporan.items():
        # Team section header
        ws.merge_cells(f'A{current_row}:H{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = f"TEAM: {team_name.upper()}"
        cell.font = section_font
        cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
        current_row += 2

        # Headers
        headers = ["No Document", "SPBU", "Lokasi", "Tanggal", "Jenis Kegiatan", "Remark", "Foto Utama", "Foto Tambahan"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = border_style

        current_row += 1

        for laporan in team_laporan:
            kegiatan_list = laporan.kegiatan_list.all()
            
            if kegiatan_list:
                for kegiatan in kegiatan_list:
                    # No Document
                    cell = ws.cell(row=current_row, column=1, value=laporan.no_document)
                    cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
                    cell.border = border_style
                    
                    # SPBU
                    spbu_text = f"{laporan.spbu.kode} - {laporan.spbu.nama}" if laporan.spbu else "Tidak ada"
                    cell = ws.cell(row=current_row, column=2, value=spbu_text)
                    cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
                    cell.border = border_style
                    
                    # Lokasi
                    loc_cell = ws.cell(row=current_row, column=3, value=laporan.lokasi)
                    loc_cell.hyperlink = f"https://www.google.com/maps/search/?api=1&query={urllib.parse.quote(laporan.lokasi)}"
                    loc_cell.style = "Hyperlink"
                    loc_cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
                    loc_cell.border = border_style

                    # Tanggal
                    cell = ws.cell(row=current_row, column=4, value=laporan.tanggal_proses.strftime("%d-%m-%Y %H:%M"))
                    cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
                    cell.border = border_style
                    
                    # Kegiatan
                    cell = ws.cell(row=current_row, column=5, value=kegiatan.get_kegiatan_display_name())
                    cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
                    cell.border = border_style
                    
                    # Remark
                    cell = ws.cell(row=current_row, column=6, value=kegiatan.remark)
                    cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
                    cell.border = border_style
                    
                    # Set row height untuk gambar dulu
                    ws.row_dimensions[current_row].height = 140

                    # Foto utama (kolom G)
                    if kegiatan.foto:
                        img_path = os.path.join(settings.MEDIA_ROOT, str(kegiatan.foto))
                        if os.path.exists(img_path):
                            img = XLImage(img_path)
                            img.width = 250
                            img.height = 180
                            img.anchor = f"G{current_row}"
                            ws.add_image(img)
                    
                    cell = ws.cell(row=current_row, column=7)
                    cell.border = border_style
                    cell.alignment = Alignment(vertical="center", horizontal="center")

                    # Foto tambahan (kolom H dan seterusnya)
                    foto_col = 8
                    for foto in kegiatan.foto_list.all():
                        img_path = os.path.join(settings.MEDIA_ROOT, str(foto.foto))
                        if os.path.exists(img_path):
                            img = XLImage(img_path)
                            img.width = 250
                            img.height = 180
                            col_letter = get_column_letter(foto_col)
                            img.anchor = f"{col_letter}{current_row}"
                            ws.add_image(img)
                            ws.column_dimensions[col_letter].width = 38
                            foto_col += 1
                    
                    cell = ws.cell(row=current_row, column=8)
                    cell.border = border_style
                    cell.alignment = Alignment(vertical="center", horizontal="center")

                    current_row += 1
            else:
                # Laporan without kegiatan
                cell = ws.cell(row=current_row, column=1, value=laporan.no_document)
                cell.alignment = Alignment(vertical="center", horizontal="center")
                cell.border = border_style

                spbu_text = f"{laporan.spbu.kode} - {laporan.spbu.nama}" if laporan.spbu else "Tidak ada"
                cell = ws.cell(row=current_row, column=2, value=spbu_text)
                cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
                cell.border = border_style

                loc_cell = ws.cell(row=current_row, column=3, value=laporan.lokasi)
                loc_cell.hyperlink = f"https://www.google.com/maps/search/?api=1&query={urllib.parse.quote(laporan.lokasi)}"
                loc_cell.style = "Hyperlink"
                loc_cell.alignment = Alignment(vertical="center", horizontal="center")
                loc_cell.border = border_style

                cell = ws.cell(row=current_row, column=4, value=laporan.tanggal_proses.strftime("%d-%m-%Y %H:%M"))
                cell.alignment = Alignment(vertical="center", horizontal="center")
                cell.border = border_style
                
                cell = ws.cell(row=current_row, column=5, value="Tidak ada kegiatan")
                cell.alignment = Alignment(vertical="center", horizontal="center")
                cell.border = border_style
                
                for col in range(6, 9):
                    cell = ws.cell(row=current_row, column=col)
                    cell.border = border_style
                
                current_row += 1

            current_row += 1  # Space between laporan

        current_row += 2  # Space between teams

    # Adjust column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 38
    ws.column_dimensions['H'].width = 38

    # Generate filename
    today_str = datetime.now().strftime("%Y-%m-%d")
    filename = f"laporan_pertareport_{today_str}.xlsx"

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def bulk_download_pdf(request):
    user = request.user

    # Ambil queryset sesuai user
    if user.username == 'admin' and user.check_password('Mimin1234%'):
        laporan_queryset = Laporan.objects.all()
    else:
        laporan_queryset = Laporan.objects.filter(
            nama_team_support__iexact=user.username
        )

    # Filter tanggal
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if start_date:
        try:
            sd = parse_date(start_date)
            if sd:
                laporan_queryset = laporan_queryset.filter(tanggal_proses__date__gte=sd)
        except ValueError:
            pass
    if end_date:
        try:
            ed = parse_date(end_date)
            if ed:
                laporan_queryset = laporan_queryset.filter(tanggal_proses__date__lte=ed)
        except ValueError:
            pass

    laporan_list = laporan_queryset.order_by('nama_team_support', '-tanggal_proses')
    if not laporan_list:
        return HttpResponse("Tidak ada laporan untuk diunduh")

    # Nama file otomatis
    today_str = timezone.now().strftime('%Y-%m-%d')
    filename = f"laporan_pertareport_{today_str}.pdf"

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    p = canvas.Canvas(response, pagesize=A4)
    page_width, page_height = A4

    # ====== Halaman Judul ======
    p.setFont("Helvetica-Bold", 24)
    p.setFillColor(HexColor('#E31E24'))
    title1 = "LAPORAN KEGIATAN"
    title2 = "PERTAMINA"
    p.drawCentredString(page_width/2, 750, title1)
    p.drawCentredString(page_width/2, 720, title2)

    # Periode
    p.setFont("Helvetica-Bold", 14)
    p.setFillColor(HexColor('#003876'))
    if start_date and end_date:
        p.drawCentredString(page_width/2, 670, f"Periode: {start_date} s/d {end_date}")
    elif start_date:
        p.drawCentredString(page_width/2, 670, f"Dari Tanggal: {start_date}")
    elif end_date:
        p.drawCentredString(page_width/2, 670, f"Sampai Tanggal: {end_date}")

    # Generated on
    p.setFont("Helvetica", 12)
    p.setFillColor(HexColor('#000000'))
    generated_text = f"Generated on: {timezone.now().strftime('%d-%m-%Y %H:%M')}"
    p.drawCentredString(page_width/2, 100, generated_text)

    # Logo
    logo_path = os.path.join(settings.STATIC_ROOT or settings.BASE_DIR, 'static', 'images', 'logo.png')
    if os.path.exists(logo_path):
        logo_width = 220
        logo_height = 150
        mid_y = (670 + 100) / 2
        p.drawImage(
            logo_path,
            (page_width - logo_width) / 2,
            mid_y - (logo_height / 2),
            width=logo_width,
            height=logo_height,
            preserveAspectRatio=True,
            mask='auto'
        )

    p.showPage()

    # ====== Isi Laporan per Team ======
    grouped_laporan = {}
    for lap in laporan_list:
        grouped_laporan.setdefault(lap.nama_team_support, []).append(lap)

    first_team = True
    for team_name, team_laporan in grouped_laporan.items():
        # Halaman baru untuk setiap team (kecuali team pertama)
        if not first_team:
            p.showPage()
        first_team = False
        
        # Header Team
        y = 800
        p.setFillColor(HexColor('#003876'))
        p.rect(40, y-5, page_width-80, 25, fill=1, stroke=0)
        p.setFont("Helvetica-Bold", 18)
        p.setFillColor(HexColor('#FFFFFF'))
        p.drawCentredString(page_width/2, y, f"TEAM: {team_name.upper()}")
        y -= 40

        for lap in team_laporan:
            # Cek apakah perlu halaman baru untuk laporan ini
            if y < 250:
                p.showPage()
                y = 800

            # Separator line
            p.setStrokeColor(HexColor('#E31E24'))
            p.setLineWidth(2)
            p.line(50, y, page_width - 50, y)
            y -= 22

            # Laporan header
            p.setFont("Helvetica-Bold", 14)
            p.setFillColor(HexColor('#E31E24'))
            p.drawString(50, y, f"Laporan: {lap.no_document}")
            y -= 22

            # Info dasar
            p.setFont("Helvetica-Bold", 10)
            p.setFillColor(HexColor('#003876'))
            p.drawString(50, y, "Lokasi:")
            p.setFont("Helvetica", 10)
            p.setFillColor(HexColor('#0066CC'))
            loc_text = lap.lokasi or ""
            # Buat lokasi jadi link yang bisa diklik
            p.linkURL(f"https://www.google.com/maps/search/?api=1&query={urllib.parse.quote(loc_text)}", 
                    (130, y-2, 500, y+12))
            p.drawString(130, y, loc_text)
            y -= 16

            p.setFont("Helvetica-Bold", 10)
            p.setFillColor(HexColor('#003876'))
            p.drawString(50, y, "SPBU:")
            p.setFont("Helvetica", 10)
            p.setFillColor(HexColor('#000000'))
            if lap.spbu:
                spbu_text = f"{lap.spbu.kode} - {lap.spbu.nama}"
                p.drawString(130, y, spbu_text)
            else:
                p.drawString(130, y, "Tidak ada")
            y -= 16

            p.setFont("Helvetica-Bold", 10)
            p.setFillColor(HexColor('#003876'))
            p.drawString(50, y, "Tanggal:")
            p.setFont("Helvetica", 10)
            p.setFillColor(HexColor('#000000'))
            p.drawString(130, y, lap.tanggal_proses.strftime('%d-%m-%Y %H:%M'))
            y -= 28

            kegiatan_list = lap.kegiatan_list.all()
            if kegiatan_list:
                for keg in kegiatan_list:
                    # Cek space untuk kegiatan + minimal 1 foto
                    if y < 280:
                        p.showPage()
                        y = 800

                    # Background untuk kegiatan
                    p.setFillColor(HexColor('#F0F4F8'))
                    p.rect(55, y-15, page_width-110, 20, fill=1, stroke=0)

                    # Kegiatan name
                    p.setFont("Helvetica-Bold", 11)
                    p.setFillColor(HexColor('#003876'))
                    p.drawString(60, y-10, f"• {keg.get_kegiatan_display_name()}")
                    y -= 22

                    # Remark
                    p.setFont("Helvetica", 10)
                    p.setFillColor(HexColor('#000000'))
                    remark_text = f"Remark: {keg.remark}"
                    p.drawString(70, y, remark_text)
                    y -= 22

                    # Foto utama
                    if keg.foto:
                        img_path = os.path.join(settings.MEDIA_ROOT, str(keg.foto))
                        if os.path.exists(img_path):
                            # Cek space untuk foto
                            if y < 195:
                                p.showPage()
                                y = 800
                            p.drawImage(img_path, 70, y-165, width=240, height=165, preserveAspectRatio=True, mask='auto')
                            y -= 175

                    # Foto tambahan
                    for foto in keg.foto_list.all():
                        img_path = os.path.join(settings.MEDIA_ROOT, str(foto.foto))
                        if os.path.exists(img_path):
                            # Cek space untuk setiap foto tambahan
                            if y < 195:
                                p.showPage()
                                y = 800
                            p.drawImage(img_path, 70, y-165, width=240, height=165, preserveAspectRatio=True, mask='auto')
                            y -= 175
                    
                    y -= 18
            else:
                p.setFont("Helvetica-Oblique", 10)
                p.setFillColor(HexColor('#666666'))
                p.drawString(70, y, "Tidak ada kegiatan yang tercatat")
                y -= 20
            
            y -= 30

    p.save()
    return response


@csrf_exempt
def history_list_api(request):
    user = request.user if request.user.is_authenticated else None
    
    if user and user.username == 'admin' and user.check_password('Mimin1234%'):
        laporan_list = Laporan.objects.all().order_by('-tanggal_proses')
    elif user:
        laporan_list = Laporan.objects.filter(
            nama_team_support__iexact=user.username
        ).order_by('-tanggal_proses')
    else:
        laporan_list = Laporan.objects.all().order_by('-tanggal_proses')

    data = {
        "status": "success",
        "laporan_list": [
            {
                "id": laporan.id,
                "no_document": laporan.no_document,
                "lokasi": laporan.lokasi,
                "spbu": {
                    "id": laporan.spbu.id,
                    "kode": laporan.spbu.kode,
                    "nama": laporan.spbu.nama,
                    "alamat": laporan.spbu.alamat
                } if laporan.spbu else None,
                "nama_team_support": laporan.nama_team_support,
                "tanggal_proses": laporan.tanggal_proses.strftime("%Y-%m-%d %H:%M:%S"),
                "kegiatan_list": [
                    {
                        "display_name": k.get_kegiatan_display_name(),
                        "remark": k.remark,
                        "foto": k.foto.url if k.foto else None,
                    }
                    for k in laporan.kegiatan_list.all()
                ],
            }
            for laporan in laporan_list
        ]
    }
    return JsonResponse(data, safe=False)